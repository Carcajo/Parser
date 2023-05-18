import json
import random
import requests
from bs4 import BeautifulSoup
import openpyxl
from time import sleep
import os
from fake_useragent import UserAgent

# зависимости
#     pip install fake-useragent
#     pip install openpyxl
#     pip install beautifulsoup4

ua = UserAgent()
s = requests.Session()

URL = "https://naos.ru"
HEADERS = {
    'User-Agent': ua.random,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'cross-site',
}

s.headers.update(HEADERS)

if not os.path.exists("data"):
    os.mkdir("data")


def get_html(url, params=None):
    if params is None:
        params = {}

    response = s.get(url, params=params)

    if response:
        return response.text
    print(f'Неверный статус код: {response.status_code}')
    return False


def parse_card(html: str, link: str) -> dict:
    html = BeautifulSoup(html, "lxml")
    card = html.find("div", class_="product grid")

    images = [URL + img.get("data-big-image") for img in
              card.find("div", class_="product__visual").find_all("div", class_="product__image-wrap swiper-slide")]

    title = card.find("div", class_="product__data").find("h1", class_="product__title").get_text(strip=True)
    sku = card.find("div", class_="product__data").find("div", class_="product__sku").find("span").get_text(strip=True)
    status = card.find("div", class_="product__data").find("div", class_="product__status").get_text(strip=True)
    price = card.find("div", class_="product__data").find("div", class_="product__price").get_text(strip=True)
    try:
        description = card.find("div", class_="product__dropdowns").find("div", class_="accordion__body").get_text(
            strip=True)
    except AttributeError:
        description = "Не удалось найти"

    # sleep(random.randint(5, 9))

    return {"images": images, "sku": sku, "title": title, "status": status, "price": price, "description": description,
            "link": link}


def parse(html):
    html = BeautifulSoup(html, 'lxml')

    return html.find("div", id="js-append").find_all("div", class_="good grid__item")


def save_data_table(datas, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Фотографии"
    ws["B1"] = "Название"
    ws["C1"] = "Артикул"
    ws["D1"] = "Статус"
    ws["E1"] = "Цена"
    ws["F1"] = "Описание"
    ws["G1"] = "Описание"

    for i, data in enumerate(datas, 2):
        ws[f"A{i}"] = "; ".join(data['images'])
        ws[f"B{i}"] = data['title']
        ws[f"C{i}"] = data['sku']
        ws[f"D{i}"] = data['status']
        ws[f"E{i}"] = data['price']
        ws[f"F{i}"] = data['description']
        ws[f"G{i}"] = data['link']

    wb.save("data\\" + filename)


def save_data_json(datas, filename):
    with open("data\\" + filename, "w", encoding="utf8") as file:
        json.dump(datas, file, ensure_ascii=False, indent=4)


def main():
    datas = []
    count = 1

    while True:
        print(f"Страница {count}  ...")
        html = get_html("https://naos.ru/catalog/", params={"PAGEN_10": count, "ajax": "y"})

        if datas and URL + parse(html)[-1].find("a", class_="good__title").get("href") == datas[-1]['link']:
            break

        if not os.path.exists(f"data\\page_{count}"):
            os.mkdir(f"data\\page_{count}")

        for i, card in enumerate(parse(html), 1):
            link = URL + card.find("a", class_="good__title").get("href")

            card_html = get_html(link)

            if not card_html:
                continue

            with open(f"data\\page_{count}\\card_{i}_{link.split('/')[-2]}.html", 'w', encoding="utf8") as file:
                file.write(card_html)

            datas.append(parse_card(card_html, link))

        count += 1

    save_data_table(datas, 'table.xlsx')
    save_data_json(datas, "file.json")

    print("Парсинг закончен!")

    os.startfile('data\\table.xlsx')
    os.startfile('data\\file.json')


if __name__ == '__main__':
    main()